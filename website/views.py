from django.shortcuts import render
from django.http import JsonResponse, HttpResponseNotAllowed, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.hashers import make_password
from .models import MUser, UserPassword, MProductType, MProduct, ManagerAccount, ManagerPassword, TOrderWaitConfirmHeader, TOrderWaitConfirmDetail,MMethodPayment, TOrderHeader, TOrderDetail
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
import re
import json
from django.contrib.auth.hashers import check_password
from django.db.models import Q  # Thêm import để tìm kiếm nâng cao
from .models import TOrderCart, TOrderPayment
from django.views.decorators.http import require_POST
from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.conf import settings
import os
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from decimal import Decimal, InvalidOperation
from django.db import IntegrityError
from django.shortcuts import get_object_or_404
from datetime import datetime
from django.db import transaction
from .models import MStatus, TOrderStatus, THistory, TProductFavorite
import json
import requests

from .models import *

# Trang chủ dành cho user
@csrf_exempt
def login_user(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)  # Đọc dữ liệu từ yêu cầu POST JSON
            email_or_phone = data.get('email-phone')
            password = data.get('password')

            if not email_or_phone or not password:
                return JsonResponse({'error': 'All fields are required.'}, status=400)

            # Tìm người dùng theo email hoặc số điện thoại, bỏ qua user bị xóa
            user = MUser.objects.filter(email=email_or_phone, delete_flag=False).first() or \
                   MUser.objects.filter(phone=email_or_phone, delete_flag=False).first()

            if not user:
                return JsonResponse({'error': 'User not found or has been deleted.'}, status=400)

            # Lấy mật khẩu đã hash từ DB
            user_password = UserPassword.objects.filter(user_id=user.user_id).first()
            if not user_password or not check_password(password, user_password.password_hash):
                return JsonResponse({'error': 'Invalid password.'}, status=400)

            # Lưu session đăng nhập
            request.session['user_id'] = user.user_id
            request.session['user_name'] = user.user_name  # Lưu username để hiển thị trên trang

            return JsonResponse({'success': 'Login successful!', 'redirect_url': '/user/home/'}, status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return render(request, 'user/login.html')

def register_user(request):
    if request.method == 'POST':
        try:
            username = request.POST.get('username')
            address = request.POST.get('address')
            phone = request.POST.get('phone')
            email = request.POST.get('email')
            password = request.POST.get('password')
            confirm_password = request.POST.get('confirm-password')

            # Kiểm tra dữ liệu đầu vào
            if not all([username, address, phone, email, password, confirm_password]):
                return JsonResponse({'error': 'All fields are required.'}, status=400)

            if password != confirm_password:
                return JsonResponse({'error': 'Passwords do not match.'}, status=400)

            if len(password) < 8:
                return JsonResponse({'error': 'Password must be at least 8 characters long.'}, status=400)

            # Kiểm tra định dạng email
            try:
                validate_email(email)
            except ValidationError:
                return JsonResponse({'error': 'Invalid email format.'}, status=400)

            # Kiểm tra số điện thoại hợp lệ
            if not re.match(r'^[0-9]{10,11}$', phone):
                return JsonResponse({'error': 'Phone number must be 10-11 digits.'}, status=400)

            # Kiểm tra xem email hoặc số điện thoại đã tồn tại trong hệ thống
            existing_user = MUser.objects.filter(email=email).first() or MUser.objects.filter(phone=phone).first()

            # Nếu user đã tồn tại nhưng bị xóa (delete_flag=True), cập nhật lại tài khoản
            if existing_user and existing_user.delete_flag:
                existing_user.user_name = username
                existing_user.address = address
                existing_user.delete_flag = False  # Mở lại tài khoản
                existing_user.save()

                # Cập nhật lại mật khẩu
                hashed_password = make_password(password)
                UserPassword.objects.update_or_create(
                    user_id=existing_user.user_id,
                    defaults={'password_hash': hashed_password}
                )

                return JsonResponse({'success': 'Your account has been reactivated!'}, status=200)

            # Nếu user chưa tồn tại, tạo mới tài khoản
            if not existing_user:
                hashed_password = make_password(password)

                user = MUser.objects.create(
                    user_name=username,
                    email=email,
                    phone=phone,
                    address=address
                )

                UserPassword.objects.create(
                    user_id=user.user_id,
                    password_hash=hashed_password
                )

                return JsonResponse({'success': 'User registered successfully!'}, status=200)

            return JsonResponse({'error': 'Email or phone number already exists.'}, status=400)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return render(request, 'user/register.html')

    # ✅ Xử lý request GET để tránh lỗi
def home_user(request):
    user_name = request.session.get('user_name', 'Guest')  # Lấy username, nếu không có thì mặc định 'Guest'
    return render(request, 'user/home.html', {'user_name': user_name})

def about_user(request):
    user_name = request.session.get('user_name', 'Guest')  # Lấy username, nếu không có thì mặc định 'Guest'
    return render(request, 'user/about.html', {'user_name': user_name})

from django.db.models import Q

def product_user(request):
    user_name = request.session.get('user_name', 'Guest')
    categories = MProductType.objects.filter(delete_flag=False)

    # Lấy tham số tìm kiếm và sắp xếp từ request
    search_query = request.GET.get('search', '').strip()
    sort_option = request.GET.get('sort', 'price-asc')  
    category_id = request.GET.get('category', None)  # Mặc định là None

    # Lọc sản phẩm (chỉ lấy sản phẩm có display_flag=True và delete_flag=False)
    products = MProduct.objects.filter(display_flag=True, delete_flag=False)

    # Nếu category_id hợp lệ (không phải "None"), lọc theo danh mục
    if category_id and category_id.lower() != "none":
        try:
            category_id = int(category_id)  # Đảm bảo là số nguyên
            products = products.filter(type__id=category_id)
        except ValueError:
            pass  # Nếu category không hợp lệ, bỏ qua

    # Nếu có tìm kiếm, kiểm tra tránh lỗi NULL
    if search_query:
        keywords = search_query.split()  # Chia chuỗi tìm kiếm thành từng từ
        search_condition = Q()

        for keyword in keywords:
            search_condition &= (
                Q(product_name__icontains=keyword) |
                Q(color__icontains=keyword)
            )

            # Kiểm tra nếu type không NULL trước khi tìm kiếm
            if MProduct.objects.filter(type__isnull=False).exists():
                search_condition |= Q(type__type_name__icontains=keyword)

        products = products.filter(search_condition)

    # Xử lý sắp xếp theo tùy chọn
    sort_mapping = {
        'price-asc': 'price',
        'price-desc': '-price',
        'name-asc': 'product_name',
        'name-desc': '-product_name',
    }
    
    if sort_option in sort_mapping:
        products = products.order_by(sort_mapping[sort_option])

    return render(request, 'user/product.html', {
        'user_name': user_name,
        'categories': categories,
        'products': products,
        'sort_option': sort_option,
        'search_query': search_query,
        'selected_category': category_id if category_id and category_id.lower() != "none" else "",  # Để giữ lại danh mục đã chọn
    })
def blog_user(request):
    user_name = request.session.get('user_name', 'Guest')  # Lấy username, nếu không có thì mặc định 'Guest'
    return render(request, 'user/blog.html', {'user_name': user_name})
def profile_user(request):
    user_id = request.session.get('user_id')  # Lấy user_id từ session
    if not user_id:
        return JsonResponse({'error': 'User not logged in.'}, status=401)

    try:
        user = MUser.objects.get(user_id=user_id)
        user_data = {
            'user_name': user.user_name,
            'email': user.email,
            'phone': user.phone,
            'address': user.address
        }
        return render(request, 'user/profile.html', user_data)

    except MUser.DoesNotExist:
        return JsonResponse({'error': 'User not found.'}, status=404)

@require_POST
def update_profile_user(request):
    user_id = request.session.get('user_id')

    if not user_id:
        return JsonResponse({
            'success': False,
            'message': 'User not logged in.'
        }, status=401)

    try:
        user = MUser.objects.get(user_id=user_id)

        data = json.loads(request.body)

        user_name = data.get('user_name', '').strip()
        email = data.get('email', '').strip()
        phone = data.get('phone', '').strip()

        # ===== Check email trùng =====
        email_exists = MUser.objects.filter(
            email=email,
            delete_flag=False
        ).exclude(user_id=user_id).exists()

        if email_exists:
            return JsonResponse({
                'success': False,
                'message': 'Email already exists.'
            })

        # ===== Check phone trùng =====
        phone_exists = MUser.objects.filter(
            phone=phone,
            delete_flag=False
        ).exclude(user_id=user_id).exists()

        if phone_exists:
            return JsonResponse({
                'success': False,
                'message': 'Phone number already exists.'
            })

        # ===== Update =====
        user.user_name = user_name
        user.email = email
        user.phone = phone
        user.save()

        # cập nhật session luôn
        request.session['user_name'] = user_name

        return JsonResponse({
            'success': True,
            'message': 'Account updated successfully.'
        })

    except MUser.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'User not found.'
        }, status=404)

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)
        
def cart_user(request):
    user_name = request.session.get('user_name', 'Guest')
    user_id = request.session.get('user_id')  # Lấy user_id từ session

    if not user_id:
        return render(request, 'user/login.html', {'message': 'Vui lòng đăng nhập để xem giỏ hàng.'})

    # Lấy các sản phẩm trong giỏ hàng của user
    cart_items = TOrderCart.objects.filter(user_id=user_id, delete_flag=False)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        # Nếu là yêu cầu AJAX, trả về giỏ hàng dưới dạng JSON
        cart_data = [{
            'product_name': item.product.name,
            'product_price': item.product.price,
            'order_cart_id': item.id
        } for item in cart_items]
        return JsonResponse({'cart_items': cart_data})

    # ➕ Tính sẵn total_price và kiểm tra sản phẩm đã bị xóa chưa
    for item in cart_items:
        item.total_price = item.product.price * item.product_quality_cart
        item.is_product_deleted = item.product.delete_flag  # Cờ để xử lý giao diện

    # ➕ Truyền thông tin vào template
    return render(request, 'user/cart.html', {
        'user_name': user_name,
        'cart_items': cart_items
    })

def delete_cart_item(request, cart_id):
    if request.method == "DELETE":
        try:
            cart_item = TOrderCart.objects.get(order_cart_id=cart_id)

            cart_item.delete()
            return JsonResponse({"success": True})
        except TOrderCart.DoesNotExist:
            return JsonResponse({"error": "Không tìm thấy sản phẩm trong giỏ hàng"}, status=404)
    else:
        return HttpResponseNotAllowed(['DELETE'])
@require_POST
def add_to_cart(request):
    print(f"Session ID: {request.session.session_key}")
    
    user_id = request.session.get('user_id')
    print(f"user_id from session: {user_id}")

    try:
        data = json.loads(request.body)
        product_id = data.get('product_id')
        product_size = data.get('product_size')
        product_quality_cart = int(data.get('product_quality_cart'))

        if not product_id or product_id == 'null' or product_id == '':
            return JsonResponse({'error': 'Invalid product_id'}, status=400)

        try:
            product = MProduct.objects.get(product_id=int(product_id))
        except (ValueError, MProduct.DoesNotExist):
            return JsonResponse({'error': 'Product not found or invalid ID'}, status=404)

        # 🔍 Kiểm tra xem đã có sản phẩm cùng product_id và size chưa
        existing_item = TOrderCart.objects.filter(
            user_id=user_id,
            product_id=product.product_id,
            product_size=product_size,
            delete_flag=False
        ).first()

        if existing_item:
            # 🔁 Cộng dồn số lượng nếu đã tồn tại
            existing_item.product_quality_cart += product_quality_cart
            existing_item.save()
        else:
            # ➕ Tạo mới nếu chưa có
            new_cart_item = TOrderCart(
                user_id=user_id,
                product_id=product.product_id,
                product_name=product.product_name,
                description=product.description,
                product_image=product.image,
                product_size=product_size,
                product_quality_cart=product_quality_cart,
                product_price=product.price,
            )
            new_cart_item.save()

        return JsonResponse({'success': 'Product added to cart successfully!'})

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
# Trang chủ dành cho admin
def home_product_manager(request):
    manager_name = request.session.get('manager_name', 'manager')
    return render(request, 'manager/homemproduct.html', {'manager_name': manager_name})
def home_manager(request):
    manager_name = request.session.get('manager_name', 'manager')
    return render(request, 'manager/home.html', {'manager_name': manager_name})
def home_order_manager(request):
    manager_name = request.session.get('manager_name', 'manager')
    return render(request, 'manager/homemorder.html', {'manager_name': manager_name})
# Trang chủ dành cho admin/product management
def mproducttype_manager(request):
    search_query = request.GET.get('search', '').strip()

    if search_query:
        product_types = MProductType.objects.filter(
            delete_flag=False
        ).filter(
            Q(type_code__icontains=search_query) | 
            Q(type_name__iexact=search_query)
        ).order_by('type_id')
    else:
        product_types = MProductType.objects.filter(delete_flag=False).order_by('type_id')

    manager_name = request.session.get('manager_name', 'manager')

    return render(request, 'manager/mproducttype.html', {
        'product_types': product_types,
        'manager_name': manager_name
    })
def check_product_type_code(request):
    type_code = request.GET.get('type_code', '').strip()
    exists = MProductType.objects.filter(type_code=type_code, delete_flag=False).exists()
    return JsonResponse({'exists': exists})
@csrf_exempt
def add_product_type(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        print('Received data:', data)  # Debug line

        try:
            type_code = data.get('type_code')
            type_name = data.get('type_name')
            note = data.get('note')

            # Kiểm tra nếu thiếu type_code hoặc type_name
            if not type_code or not type_name:
                return JsonResponse({'success': False, 'message': 'Missing type_code or type_name.'})

            # Kiểm tra trùng lặp type_code chỉ với những bản ghi chưa bị xóa
            if MProductType.objects.filter(type_code=type_code, delete_flag=False).exists():
                return JsonResponse({'success': False, 'message': 'Type code already exists.'})

            # Nếu type_code đã bị xóa (delete_flag=True), khôi phục lại và cập nhật các trường
            existing_product_type = MProductType.objects.filter(type_code=type_code, delete_flag=True).first()
            if existing_product_type:
                # Khôi phục lại bản ghi và cập nhật các trường khác
                existing_product_type.delete_flag = False
                existing_product_type.type_name = type_name  # Cập nhật type_name
                existing_product_type.note = note  # Cập nhật note
                existing_product_type.save()
                return JsonResponse({'success': True, 'message': 'Restored and updated deleted type code.'})

            # Nếu không có bản ghi trùng lặp, tạo mới
            MProductType.objects.create(
                type_code=type_code,
                type_name=type_name,
                note=note
            )
            return JsonResponse({'success': True})
        except Exception as e:
            print('Error:', str(e))  # Debug line
            return JsonResponse({'success': False, 'message': str(e)})

@csrf_exempt
def delete_product_types(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        type_codes = data.get('type_codes', [])

        for code in type_codes:
            MProductType.objects.filter(type_code=code, delete_flag=False).update(delete_flag=True)

        return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'message': 'Invalid request.'})
@csrf_exempt
def update_product_types(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        updates = data.get('updates', [])

        for item in updates:
            original_code = item.get('original_type_code')
            new_code = item.get('type_code')
            new_name = item.get('type_name')
            note = item.get('note')

            # Nếu type_code bị đổi, cần check xem đã tồn tại chưa
            if original_code != new_code:
                if MProductType.objects.filter(type_code=new_code, delete_flag=False).exists():
                    return JsonResponse({'success': False, 'message': f'Type code {new_code} already exists.'})

            # Update dữ liệu
            product_type = MProductType.objects.get(type_code=original_code, delete_flag=False)
            product_type.type_code = new_code
            product_type.type_name = new_name
            product_type.note = note
            product_type.save()

        return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'message': 'Invalid request.'})

def login_manager(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email_or_phone = data.get('email-phone')
            password = data.get('password')

            if not email_or_phone or not password:
                return JsonResponse({'error': 'All fields are required.'}, status=400)

            # Tìm kiếm quản lý theo email hoặc số điện thoại
            manager = ManagerAccount.objects.filter(email=email_or_phone).first() or \
                     ManagerAccount.objects.filter(phone=email_or_phone).first()

            if not manager:
                return JsonResponse({'error': 'Manager not found or has been deleted.'}, status=400)

            # Kiểm tra mật khẩu
            manager_password = ManagerPassword.objects.filter(user=manager).first()  # Sửa ở đây
            if not manager_password or not check_password(password, manager_password.password_hash):
                return JsonResponse({'error': 'Invalid password.'}, status=400)

            # Lưu thông tin vào session (sửa id thành user_id)
            request.session['manager_id'] = manager.user_id  # Dùng user_id thay vì id
            request.session['manager_name'] = manager.user_name  # Lưu tên để hiển thị

            return JsonResponse({'success': 'Login successful!', 'redirect_url': '/manager/home/'}, status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return render(request, 'manager/login.html')
from django.db.models import Q

from django.db.models import Q

def mproduct_manager(request):
    manager_name = request.session.get('manager_name', 'manager')

    product_types = MProductType.objects.filter(delete_flag=False)

    search_query = request.GET.get('search', '').strip()
    type_filter = request.GET.get('type_filter', '').strip()

    products = MProduct.objects.filter(
        delete_flag=False
    )

    # Search
    if search_query:
        products = products.filter(
            Q(product_code__icontains=search_query) |
            Q(product_name__icontains=search_query)
        )

    # Filter theo Product Type
    if type_filter:
        products = products.filter(
            type_code=type_filter
        )

    products = products.order_by('product_id')

    return render(request, 'manager/mproduct.html', {
        'manager_name': manager_name,
        'products': products,
        'product_types': product_types,
    })
def add_product(request):
    if request.method == 'POST':
        product_code = request.POST.get('product_code')
        product_name = request.POST.get('product_name')
        type_code = request.POST.get('type_code')
        image_file = request.FILES.get('image')
        description = request.POST.get('description')[:255]  # giới hạn 255 ký tự
        price = request.POST.get('price')
        color = request.POST.get('color')
        display_flag = 'display_flag' in request.POST
        is_freesize = 'is_freesize' in request.POST

        size_s = int(request.POST.get('size_s') or 0)
        size_m = int(request.POST.get('size_m') or 0)
        size_l = int(request.POST.get('size_l') or 0)
        size_freesize = int(request.POST.get('size_freesize') or 0)

        inventory_total = size_s + size_m + size_l + size_freesize

        # Kiểm tra trùng product_code với điều kiện delete_flag = False
        if MProduct.objects.filter(product_code=product_code, delete_flag=False).exists():
            messages.error(request, 'Product code already exists (active product).')
            return redirect('/manager/mproduct')  # Hoặc render lại trang với thông báo lỗi

        # Tìm type_id từ type_code
        try:
            product_type = MProductType.objects.get(type_code=type_code)
        except MProductType.DoesNotExist:
            messages.error(request, 'Invalid product type.')
            return redirect('/manager/mproduct')

        # Xử lý upload ảnh
        image_url = ''
        if image_file:
            fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'product_images'))
            filename = fs.save(image_file.name, image_file)
            image_url = fs.url('product_images/' + filename)

        # Tạo sản phẩm mới
        MProduct.objects.create(
            product_code=product_code,
            product_name=product_name,
            type_id=product_type.type_id,
            type_code=type_code,
            image=image_url,
            description=description,
            price=price,
            color=color,
            display_flag=display_flag,
            is_freesize=is_freesize,
            quality_inventory_of_size_S=size_s,
            quality_inventory_of_size_M=size_m,
            quality_inventory_of_size_L=size_l,
            quality_inventory_of_freesize=size_freesize,
            quality_inventory_total=inventory_total,
            delete_flag=False
        )

        messages.success(request, 'Product added successfully.')
        return redirect('/manager/mproduct')

    # Nếu GET thì render form để thêm sản phẩm
    return render(request, '/manager/mproduct.html')
def get_product_details(request, product_id):
    try:
        product = MProduct.objects.get(product_id=product_id)

        # Lấy đường dẫn hình ảnh từ CSDL
        image_path = product.image or ''
        image_name = ''

        if image_path:
            # Kiểm tra xem đường dẫn có phải là một tệp hợp lệ không
            # Nếu tệp có tồn tại trong thư mục media không
            full_image_path = os.path.join(settings.MEDIA_ROOT, image_path.strip('/'))

            # Kiểm tra xem tệp có tồn tại không
            if os.path.isfile(full_image_path):
                image_name = os.path.basename(image_path)  # Chỉ lấy tên tệp
            else:
                image_name = 'no_image_available.png'  # Hoặc bất kỳ tên tệp nào thay thế khi không tìm thấy

        data = {
            'product_code': product.product_code,
            'product_name': product.product_name,
            'type_code': product.type_code,
            'image': image_name,  # Gửi tên hình ảnh thay vì URL đầy đủ
            'description': product.description,
            'price': product.price,
            'color': product.color,
            'display_flag': product.display_flag,
            'is_freesize': product.is_freesize,
            'size_s': product.quality_inventory_of_size_S,
            'size_m': product.quality_inventory_of_size_M,
            'size_l': product.quality_inventory_of_size_L,
            'size_freesize': product.quality_inventory_of_freesize,
            'inventory_total': product.quality_inventory_total
        }

        return JsonResponse({'success': True, 'product': data})
    except MProduct.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Product not found'})
@csrf_exempt
def delete_selected_products(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            product_ids = data.get('product_ids', [])

            if not product_ids:
                return JsonResponse({'success': False, 'error': 'No products selected.'})

            # Cập nhật cờ delete_flag = True
            MProduct.objects.filter(product_id__in=product_ids).update(delete_flag=True)

            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})
@csrf_exempt
def edit_product(request):
    if request.method == 'POST':
        product_code = request.POST.get('product_code')  # Khóa chính không đổi
        product_name = request.POST.get('product_name')
        type_code = request.POST.get('type_code')
        image_file = request.FILES.get('image')
        description = request.POST.get('description')[:255]
        price = request.POST.get('price')
        color = request.POST.get('color')
        display_flag = 'display_flag' in request.POST
        is_freesize = 'is_freesize' in request.POST

        size_s = int(float(request.POST.get('size_s') or 0))
        size_m = int(float(request.POST.get('size_m') or 0))
        size_l = int(float(request.POST.get('size_l') or 0))
        size_freesize = int(float(request.POST.get('size_freesize') or 0))

        inventory_total = size_s + size_m + size_l + size_freesize

        try:
            product = MProduct.objects.get(product_code=product_code, delete_flag=False)
        except MProduct.DoesNotExist:
            messages.error(request, 'Product not found.')
            return redirect('/manager/mproduct')

        try:
            product_type = MProductType.objects.get(type_code=type_code)
        except MProductType.DoesNotExist:
            messages.error(request, 'Invalid product type.')
            return redirect('/manager/mproduct')

        product.product_name = product_name
        product.type_id = product_type.type_id
        product.type_code = type_code
        product.description = description
        product.price = price
        product.color = color
        product.display_flag = display_flag
        product.is_freesize = is_freesize
        product.quality_inventory_of_size_S = size_s
        product.quality_inventory_of_size_M = size_m
        product.quality_inventory_of_size_L = size_l
        product.quality_inventory_of_freesize = size_freesize
        product.quality_inventory_total = inventory_total

        # Nếu người dùng upload ảnh mới
        if image_file:
            fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'product_images'))
            filename = fs.save(image_file.name, image_file)
            product.image = fs.url('product_images/' + filename)

        product.save()
        messages.success(request, 'Product updated successfully.')
        return redirect('/manager/mproduct')

    return redirect('/manager/mproduct')
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Protection

def export_products_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = [
        "Product Code",
        "Product Name",
        "Description",
        "Price",
        "Display Flag",
        "Size S",
        "Size M",
        "Size L",
        "Size Freesize",
    ]

    ws.append(headers)

    products = MProduct.objects.filter(delete_flag=False)

    for p in products:
        ws.append([
            p.product_code,
            p.product_name,
            p.description,
            p.price,
            p.display_flag,
            p.quality_inventory_of_size_S,
            p.quality_inventory_of_size_M,
            p.quality_inventory_of_size_L,
            p.quality_inventory_of_freesize,
        ])

    # Khóa cột Product Code
    for row in ws.iter_rows(min_row=2):
        row[0].protection = Protection(locked=True)

        for cell in row[1:]:
            cell.protection = Protection(locked=False)

    ws.protection.sheet = True

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = (
        'attachment; filename=products.xlsx'
    )

    wb.save(response)

    return response
from openpyxl import load_workbook
from django.contrib import messages
from django.shortcuts import redirect

def import_products_excel(request):

    if request.method == "POST":

        excel_file = request.FILES["excel_file"]

        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):

            product_code = row[0]

            try:
                product = MProduct.objects.get(
                    product_code=product_code,
                    delete_flag=False
                )

                product.product_name = row[1] or ""
                product.description = row[2] or ""
                product.price = row[3] or 0
                product.display_flag = bool(row[4])

                size_s = row[5] or 0
                size_m = row[6] or 0
                size_l = row[7] or 0
                size_freesize = row[8] or 0

                product.quality_inventory_of_size_S = size_s
                product.quality_inventory_of_size_M = size_m
                product.quality_inventory_of_size_L = size_l
                product.quality_inventory_of_freesize = size_freesize

                product.quality_inventory_total = (
                    size_s +
                    size_m +
                    size_l +
                    size_freesize
                )

                product.save()

            except MProduct.DoesNotExist:
                continue

        messages.success(
            request,
            "Import products successfully."
        )

    return redirect('/manager/mproduct')
@csrf_exempt
def checkout_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            order_cart_id = data.get('order_cart_id', [])

            if not order_cart_id:
                return JsonResponse({'success': False, 'message': 'Không có sản phẩm nào được chọn.'})

            # Lấy danh sách cart
            cart_items = TOrderCart.objects.filter(order_cart_id__in=order_cart_id, delete_flag=False)

            if not cart_items:
                return JsonResponse({'success': False, 'message': 'Giỏ hàng không có sản phẩm hợp lệ.'})

            # Kiểm tra sản phẩm nào đã bị xóa hoặc hết hàng
            invalid_items = []
            for item in cart_items:
                if item.product.delete_flag:
                    invalid_items.append(item.product.product_name)
                elif item.product.quality_inventory_total == 0:
                    invalid_items.append(f"{item.product.product_name} (Sold Out)")

            if invalid_items:
                return JsonResponse({
                    'success': False,
                    'message': f"Products invalid: {', '.join(invalid_items)}. Please check your cart again."
                })

            # Tới đây là OK → chuyển sang trang thanh toán
            return JsonResponse({
                'success': True,
                'redirect_url': '/user/payment/'
            })

        except Exception as e:
            print(f"Error: {str(e)}")  # In lỗi ra console hoặc log
            return JsonResponse({'success': False, 'message': 'Có lỗi xảy ra. Vui lòng thử lại.'})
def payment_user(request):
    user_name = request.session.get('user_name', 'Guest')  # Lấy username, nếu không có thì mặc định 'Guest'
    return render(request, 'user/payment.html', {'user_name': user_name})
@csrf_exempt
def checkout_details_view(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            cart_items = data.get('cart_items', [])

            for item in cart_items:
                cart_id = item.get('order_cart_id')
                size = item.get('product_size')
                quantity = int(item.get('product_quantity'))
                price = float(item.get('product_price'))

                try:
                    # Thêm điều kiện delete_flag = False
                    cart = TOrderCart.objects.get(order_cart_id=cart_id, delete_flag=False)
                except TOrderCart.DoesNotExist:
                    continue  # Bỏ qua nếu sản phẩm đã bị xóa mềm

                TOrderPayment.objects.create(
                    user=cart.user,
                    product=cart.product,
                    product_name=cart.product_name,
                    description=cart.description,
                    product_image=cart.product_image,
                    product_size=size or cart.product_size,
                    product_quality_payment=quantity or cart.product_quality_cart,
                    product_price=price or cart.product_price
                )

            return JsonResponse({"success": True})
        except Exception as e:
            print("Error saving detailed checkout:", e)
            return JsonResponse({"success": False, "message": str(e)})

    return JsonResponse({"success": False, "message": "Invalid request method."})
def payment_user(request):
    user_name = request.session.get('user_name', 'Guest')
    user_id = request.session.get('user_id')

    if not user_id:
        return render(request, 'user/login.html', {'message': 'Vui lòng đăng nhập để xem giỏ hàng.'})

    # Truy vấn địa chỉ của user từ bảng MUser
    try:
        user = MUser.objects.get(user_id=user_id, delete_flag=False)
        user_address = user.address
    except MUser.DoesNotExist:
        user_address = "Không có địa chỉ"

    orders = TOrderPayment.objects.filter(user_id=user_id, delete_flag=False)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        cart_data = [{
            'product_name': item.product.name,
            'product_price': item.product.price,
            'order_payment_id': item.id
        } for item in orders]
        return JsonResponse({'cart_items': cart_data})

    return render(request, 'user/payment.html', {
        'user_name': user_name,
        'orders': orders,
        'user_address': user_address  # ➕ truyền địa chỉ đến template
    })
@csrf_exempt
def create_order(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ'}, status=400)

    user_id = request.session.get('user_id')
    if not user_id:
        return JsonResponse({'success': False, 'message': 'Vui lòng đăng nhập để tạo đơn hàng.'}, status=401)

    try:
        m_user = MUser.objects.get(user_id=user_id, delete_flag=False)
    except MUser.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Người dùng không tồn tại.'}, status=404)

    total_price_order = request.POST.get('total_price_order')
    payment_method_id = request.POST.get('payment_method_id')
    note = request.POST.get('note', '')

    try:
        payment_method = MMethodPayment.objects.get(pk=payment_method_id)
    except MMethodPayment.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Phương thức thanh toán không hợp lệ.'}, status=400)

    order = TOrderWaitConfirmHeader.objects.create(
        user=m_user,
        user_name=m_user.user_name,
        order_date=timezone.now(),
        total_price_order=total_price_order,
        payment_method=payment_method,
        delete_flag=False,
        estimated_delivery_date=timezone.now() + timezone.timedelta(days=3),
        note=note
    )

    try:
        products_data = json.loads(request.POST.get('products', '[]'))
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Dữ liệu sản phẩm không hợp lệ.'}, status=400)

    for p in products_data:
        try:
            product = MProduct.objects.get(pk=p['product_id'], delete_flag=False)
        except MProduct.DoesNotExist:
            continue

        try:
            quantity = Decimal(str(p.get('quantity', '1')).replace(',', '.'))
        except InvalidOperation:
            quantity = Decimal('1')

        try:
            # ⚡ Kiểm tra xem chi tiết sản phẩm đã tồn tại chưa
            existing_detail = TOrderWaitConfirmDetail.objects.filter(
                order=order,
                product=product,
                product_size=p.get('product_size', '')
            ).first()

            if existing_detail:
                # Nếu đã có thì cộng dồn số lượng
                existing_detail.quality += quantity
                existing_detail.save()
            else:
                # Nếu chưa có thì tạo mới
                TOrderWaitConfirmDetail.objects.create(
                    order=order,
                    product=product,
                    product_size=p.get('product_size', ''),
                    type=product.type,
                    quality=quantity,
                    price_product=p.get('price_product', ''),
                    delete_flag=False
                )
        except IntegrityError as e:
            print(f"Lỗi thêm chi tiết đơn hàng: {e}")
            continue
        TOrderPayment.objects.filter(
    user=m_user,
    delete_flag=False
).update(delete_flag=True)
    return JsonResponse({'success': True, 'order_id': order.order_id})
def order_wait_confirm_management(request):
    manager_name = request.session.get('manager_name', 'manager')

    # 👉 Lọc chỉ các đơn chưa bị xóa
    orders = TOrderWaitConfirmHeader.objects.filter(delete_flag=False).select_related('user')

    return render(request, 'manager/morderwait.html', {
        'manager_name': manager_name,
        'orders': orders
    })
   
def order_wait_confirm_detail(request, order_id):

    manager_name = request.session.get('manager_name', 'manager')

    # =========================
    # Lấy thông tin đơn hàng
    # =========================
    header = (
        TOrderWaitConfirmHeader.objects
        .select_related('user', 'payment_method')
        .filter(
            order_id=order_id,
            delete_flag=False
        )
        .first()
    )

    if not header:
        return render(
            request,
            'manager/orderwaitdetail.html',
            {
                'manager_name': manager_name,
                'error_message': 'Không tìm thấy đơn hàng.'
            }
        )

    # =========================
    # Lấy chi tiết đơn hàng
    # =========================
    details = (
        TOrderWaitConfirmDetail.objects
        .filter(order_id=order_id)
        .select_related(
            'product',
            'type'
        )
    )

    product_list = []

    for idx, item in enumerate(details, start=1):

        inventory = 0
        size = ""

        if item.product:

            size = str(item.product_size).strip().upper()

# bỏ chữ "SIZE:"
            size = size.replace("SIZE:", "").strip()

            if size == "S":
                inventory = item.product.quality_inventory_of_size_S

            elif size == "M":
                inventory = item.product.quality_inventory_of_size_M

            elif size == "L":
                inventory = item.product.quality_inventory_of_size_L

            elif size in ["FREE", "FREESIZE", "FREE SIZE"]:
                inventory = item.product.quality_inventory_of_freesize

        # ===== DEBUG =====
        print("\n====================")
        print("Order ID:", order_id)
        print("Product ID:", item.product.product_id)
        print("Product Name:", item.product.product_name)
        print("Size:", size)

        print(
            "Stock S:",
            item.product.quality_inventory_of_size_S
        )

        print(
            "Stock M:",
            item.product.quality_inventory_of_size_M
        )

        print(
            "Stock L:",
            item.product.quality_inventory_of_size_L
        )

        print(
            "Stock Free:",
            item.product.quality_inventory_of_freesize
        )

        print("Inventory Result:", inventory)
        print("====================\n")
        # =================

        product_list.append({
            'no': idx,
            'product_code': item.product.product_code,
            'product_name': item.product.product_name,
            'product_id': item.product.product_id,
            'type_id': item.type.type_id if item.type else None,
            'type_code': item.type.type_code if item.type else '',
            'product_size': item.product_size,
            'quantity': item.quality,
            'inventory': inventory,
            'price': item.price_product
        })

    # =========================
    # Format tổng tiền
    # =========================
    try:
        total_price_formatted = (
            f"{float(header.total_price_order):,.0f} VNĐ"
        )
    except (ValueError, TypeError):
        total_price_formatted = header.total_price_order

    payment_method_name = (
        header.payment_method.method_name
        if header.payment_method
        else "Không xác định"
    )

    payment_method_id = (
        header.payment_method.method_id
        if header.payment_method
        else None
    )

    context = {
        'manager_name': manager_name,
        'order_id': header.order_id,
        'user_id': header.user.user_id,
        'user_name': header.user.user_name,
        'order_date': header.order_date,
        'estimated_delivery_date': header.estimated_delivery_date,
        'total_price_order': total_price_formatted,
        'payment_method_name': payment_method_name,
        'payment_method_id': payment_method_id,
        'note': header.note or '',
        'product_list': product_list,
        'user_phone': header.user.phone,
    }

    return render(
        request,
        'manager/orderwaitdetail.html',
        context
    )
@csrf_exempt
def create_order_header(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Phương thức không hợp lệ"}, status=405)

    try:
        data = json.loads(request.body)

        user_id = data.get("user_id")
        user_name = data.get("user_name")
        order_date = data.get("order_date")
        estimated_delivery_date = data.get("estimated_delivery_date")
        total_price_order = data.get("total_price_order")
        payment_method_id = data.get("payment_method_id")
        note = data.get("note", "")
        product_details = data.get("product_details", [])
        order_wait_id = data.get("order_id")  # 🧩 lấy order_id từ frontend

        # 🧩 Kiểm tra user và phương thức thanh toán
        user = MUser.objects.get(user_id=user_id)
        payment_method = MMethodPayment.objects.get(method_id=payment_method_id)

        # 🧩 Convert ngày
        order_date_obj = datetime.strptime(order_date, "%Y-%m-%d").date()
        estimated_date_obj = datetime.strptime(estimated_delivery_date, "%Y-%m-%d").date()

        # 🧩 Tạo record header
        new_order = TOrderHeader.objects.create(
            user=user,
            user_name=user_name,
            order_date=order_date_obj,
            estimated_delivery_date=estimated_date_obj,
            total_price_order=Decimal(total_price_order),
            payment_method=payment_method,
            status_id="01",          
            note=note,
            delete_flag=False
        )

        # 🧩 Tạo các record detail
        for item in product_details:
            product = MProduct.objects.get(product_id=item["product_id"])
            type_obj = None
            if item.get("type_id"):
                type_obj = MProductType.objects.get(type_id=item["type_id"])

            TOrderDetail.objects.create(
                order=new_order,
                product=product,
                product_size=item["product_size"],
                type=type_obj,
                quality=Decimal(item["quality"]),
                price_product=Decimal(item["price_product"]),
                delete_flag=False,
            )

        # ✅ Sau khi tạo xong đơn hàng → cập nhật delete_flag trong bảng t_order_wait_confirm
        if order_wait_id:
            TOrderWaitConfirmHeader.objects.filter(order_id=order_wait_id).update(delete_flag=True)
            TOrderWaitConfirmDetail.objects.filter(order_id=order_wait_id).update(delete_flag=True)

        return JsonResponse({
            "success": True,
            "message": f"Order #{new_order.order_id} has been created successfully"
        })

    except Exception as e:
        return JsonResponse({"success": False, "message": f"Lỗi: {str(e)}"}, status=500)




def user_manage(request):
    manager_name = request.session.get('manager_name', 'manager')
    
    # Lấy tất cả user, kể cả đã bị xóa (delete_flag=True)
    users = MUser.objects.all().order_by('user_id')
    
    # Gửi dữ liệu xuống template
    return render(request, 'manager/muser.html', {
        'manager_name': manager_name,
        'users': users
    })
from django.db.models import Sum
from django.shortcuts import render, get_object_or_404

def user_manage_detail(request, user_id):
    manager_name = request.session.get('manager_name', 'manager')

    user = get_object_or_404(
        MUser,
        user_id=user_id
    )

    orders = (
        TOrderHeader.objects
        .filter(
            user=user,
            delete_flag=False
        )
        .order_by('-order_date')
    )

    total_price = (
        orders.aggregate(
            total=Sum('total_price_order')
        )['total']
        or 0
    )

    return render(
        request,
        'manager/muserdetail.html',
        {
            'manager_name': manager_name,
            'user': user,
            'orders': orders,
            'total_price': total_price
        }
    )
#--button revenue------------------------------------------------------------------------------------------------
def manage_revenue(request):
    manager_name = request.session.get('manager_name', 'manager')
    return render(request, 'manager/revenue.html', {'manager_name': manager_name})
from django.http import JsonResponse
from django.db.models import Sum

def revenue_api(request):

    from_date = request.GET.get("from")
    to_date = request.GET.get("to")

    order_ids = (
        TOrderStatus.objects.filter(
            delete_flag=False,
            new_status_id__in=["04", "05"],
            created_at__date__range=[
                from_date,
                to_date
            ]
        )
        .values_list(
            "order_id",
            flat=True
        )
        .distinct()
    )

    print("ORDER IDS:", list(order_ids))

    orders = TOrderHeader.objects.filter(
        order_id__in=order_ids,
        delete_flag=False
    )

    print("ORDERS COUNT:", orders.count())

    cart_total = (
        orders.filter(
            payment_method_id=2
        )
        .aggregate(total=Sum("total_price_order"))["total"]
        or 0
    )

    cash_total = (
        orders.filter(
            payment_method_id=1
        )
        .aggregate(total=Sum("total_price_order"))["total"]
        or 0
    )

    print("CART:", cart_total)
    print("CASH:", cash_total)

    return JsonResponse({
        "cart_total": float(cart_total),
        "cash_total": float(cash_total)
    })
from django.http import JsonResponse
from django.db.models import Count

def favorite_product_api(request):

    products = (
        TProductFavorite.objects
        .filter(
            favorite_flag=True,
            delete_flag=False
        )
        .values(
            "product__product_id"
        )
        .annotate(
            total_favorite=Count("user")
        )
        .order_by(
            "-total_favorite"
        )[:10]
    )

    labels = []
    data = []

    for item in products:
        labels.append(
            item["product__product_id"]
        )
        data.append(
            item["total_favorite"]
        )

    return JsonResponse({
        "labels": labels,
        "data": data
    })
def best_selling_product_api(request):

    products = (
        TOrderDetail.objects
        .filter(
            delete_flag=False,
            order__status_id__in=["04", "05"]
        )
        .values(
            "product__product_id",
            "product__product_name"
        )
        .annotate(
            total_sold=Sum("quality")
        )
        .order_by(
            "-total_sold"
        )[:10]
    )

    labels = []
    data = []

    for item in products:

        labels.append(
            f'{item["product__product_id"]} - {item["product__product_name"]}'
        )

        data.append(
            float(item["total_sold"])
        )

    return JsonResponse({
        "labels": labels,
        "data": data
    })
# -- button all order ---------------------------------------------------------


def all_order(request):
    manager_name = request.session.get(
        'manager_name',
        'manager'
    )

    orders = (
        TOrderHeader.objects
        .select_related(
            'user',
            'payment_method',
            'status'
        )
        .filter(delete_flag=False)
        .order_by('-order_id')
    )

    today = timezone.now().date()

    for order in orders:

        days_left = (
            order.estimated_delivery_date - today
        ).days

        order.is_urgent = (
            order.estimated_delivery_date is not None
            and 0 <= days_left <= 3
        )

        order.status_display = (
            order.status.status_name
            if order.status
            else ""
        )

    return render(
        request,
        'manager/allorder.html',
        {
            'manager_name': manager_name,
            'orders': orders
        }
    )
    #--button detail order------------------------------------------------------------------------------------------------
from django.shortcuts import render, get_object_or_404, redirect
from .models import TOrderHeader, TOrderDetail

def order_detail(request, order_id):

    manager_name = request.session.get(
        "manager_name",
        "manager"
    )

    order_header = get_object_or_404(
        TOrderHeader.objects.select_related("status"),
        order_id=order_id,
        delete_flag=False
    )

    if request.method == "POST":

        status_mapping = {
            "confirmed": "01",
            "waiting": "02",
            "picking": "03",
            "delivery": "04",
            "done": "05",
            "cancelled": "06",
        }

        status_level = {
            "01": 1,
            "02": 2,
            "03": 3,
            "04": 4,
            "05": 5,
        }

        selected_status = request.POST.get(
            "order_status"
        )

        reason = request.POST.get(
            "change_reason",
            ""
        ).strip()

        old_status = order_header.status.status_id

        new_status = status_mapping.get(
            selected_status
        )

        if not new_status:

            messages.error(
                request,
                "Invalid status."
            )

            return redirect(
                "order_detail",
                order_id=order_id
            )

        # DONE không được đổi
        if old_status == "05":

            messages.error(
                request,
                "Done order cannot be changed."
            )

            return redirect(
                "order_detail",
                order_id=order_id
            )

        # CANCELLED không được đổi
        if old_status == "06":

            messages.error(
                request,
                "Cancelled order cannot be changed."
            )

            return redirect(
                "order_detail",
                order_id=order_id
            )

        # CANCEL
        if new_status == "06":

            if not reason:

                messages.error(
                    request,
                    "Cancel reason is required."
                )

                return redirect(
                    "order_detail",
                    order_id=order_id
                )

        else:

            old_level = status_level[
                old_status
            ]

            new_level = status_level[
                new_status
            ]

            # Không cho nhảy cóc
            if abs(
                new_level - old_level
            ) > 1:

                messages.error(
                    request,
                    "Status can only be changed one step at a time."
                )

                return redirect(
                    "order_detail",
                    order_id=order_id
                )

            # Update ngược
            if (
                new_level < old_level
                and not reason
            ):

                messages.error(
                    request,
                    "Reason is required."
                )

                return redirect(
                    "order_detail",
                    order_id=order_id
                )

        # ==================================
        # UPDATE EDIT DATA
        # ==================================

        old_estimate_date = (
            order_header.estimated_delivery_date
        )

        estimated_delivery_date = request.POST.get(
            "estimated_delivery_date"
        )

        if (
            estimated_delivery_date
            and str(old_estimate_date)
            != estimated_delivery_date
        ):

            THistory.objects.create(
                order=order_header,
                user=order_header.user,
                product=None,
                current_estimate_date=old_estimate_date,
                new_estimate_date=estimated_delivery_date,
                delete_flag=False
            )

            order_header.estimated_delivery_date = (
                estimated_delivery_date
            )

        elif estimated_delivery_date:

            order_header.estimated_delivery_date = (
                estimated_delivery_date
            )

        order_details = (
            TOrderDetail.objects.filter(
                order=order_header,
                delete_flag=False
            )
        )

        total_price = 0

        for index, detail in enumerate(
            order_details,
            start=1
        ):

            old_size = detail.product_size
            old_quantity = detail.quality

            quantity = request.POST.get(
                f"quantity_{index}"
            )

            product_size = request.POST.get(
                f"product_size_{index}"
            )

            quantity_changed = (
                quantity
                and float(quantity) != old_quantity
            )

            size_changed = (
                product_size
                and product_size != old_size
            )

            if quantity_changed or size_changed:

                THistory.objects.create(
                    order=order_header,
                    user=order_header.user,
                    product=detail.product,
                    current_order_size=old_size,
                    new_order_size=(
                        product_size or old_size
                    ),
                    current_quantity=int(old_quantity),
                    new_quantity=int(float(quantity))
        if quantity
        else int(old_quantity),
                    current_estimate_date=None,
                    new_estimate_date=None,
                    delete_flag=False
                )

            if quantity:
                detail.quality = quantity

            if product_size:
                detail.product_size = (
                    product_size
                )

            detail.save()

            total_price += (
                float(detail.quality)
                * float(detail.price_product)
            )

        order_header.total_price_order = (
            total_price
        )

        # ==================================
        # SAVE HISTORY STATUS
        # ==================================

        current_status_obj = (
            order_header.status
        )

        new_status_obj = (
            MStatus.objects.get(
                status_id=new_status
            )
        )

        TOrderStatus.objects.create(
            order=order_header,
            current_status=current_status_obj,
            new_status=new_status_obj,
            note=reason,
            delete_flag=False
        )

        # ==================================
        # UPDATE HEADER
        # ==================================

        order_header.status = (
            new_status_obj
        )

        if new_status == "06":
            order_header.delete_flag = True

        order_header.save()

        messages.success(
            request,
            "Order status updated successfully."
        )

        return redirect(
            "all_order"
        )

    order_details = (
        TOrderDetail.objects
        .select_related(
            "product",
            "type"
        )
        .filter(
            order=order_header,
            delete_flag=False
        )
    )
    edited_products = list(
    THistory.objects.filter(
        order=order_header,
        product__isnull=False,
        delete_flag=False
    )
    .values_list(
        "product__product_id",
        flat=True
    )
    .distinct()
)
    edited_estimate = THistory.objects.filter(
        order=order_header,
        current_estimate_date__isnull=False,
        delete_flag=False
    ).exists()

    edited_products = list(
        THistory.objects.filter(
            order=order_header,
            product__isnull=False,
            delete_flag=False
        )
        .values_list(
            "product__product_id",
            flat=True
        )
        .distinct()
    )

    return render(
        request,
        "manager/orderdetail.html",
        {
            "manager_name": manager_name,
            "order_header": order_header,
            "order_details": order_details,
            "edited_products": edited_products,
            "edited_estimate": edited_estimate,
        }
    )
def create_order_manage(request):

    manager_name = request.session.get(
        'manager_name',
        'manager'
    )

    products = (
        MProduct.objects
        .select_related('type')
        .filter(delete_flag=False)
    )

    method_payments = (
        MMethodPayment.objects
        .filter(delete_flag=False)
        .order_by('method_name')
    )

    product_data = []

    for p in products:

        sizes = []

        if p.is_freesize:

            sizes.append({
                "size": "Free Size",
                "stock": int(
                    p.quality_inventory_of_freesize or 0
                )
            })

        else:

            sizes.append({
                "size": "S",
                "stock": int(
                    p.quality_inventory_of_size_S or 0
                )
            })

            sizes.append({
                "size": "M",
                "stock": int(
                    p.quality_inventory_of_size_M or 0
                )
            })

            sizes.append({
                "size": "L",
                "stock": int(
                    p.quality_inventory_of_size_L or 0
                )
            })

        product_data.append({
    "product_id": p.product_id,
    "product_code": p.product_code,
    "price": float(p.price or 0),
    "sizes": sizes
})

    return render(
        request,
        'manager/createorder.html',
        {
            'manager_name': manager_name,
            'products': products,
            'product_data': product_data,
            'method_payments': method_payments
        }
    )
def get_user_by_phone(request):

    phone = request.GET.get('phone')

    try:
        user = MUser.objects.get(
            phone=phone,
            delete_flag=False
        )

        return JsonResponse({
            'found': True,
            'user_name': user.user_name
        })

    except MUser.DoesNotExist:

        return JsonResponse({
            'found': False
        })
@require_POST
@transaction.atomic
def save_order(request):

    try:

        data = json.loads(request.body)

        phone = data.get("phone")
        user_name = data.get("user_name")

        order_date = data.get("order_date")
        estimated_delivery = data.get(
            "estimated_delivery"
        )

        total_price = data.get("total_price", 0)

        payment_method_id = data.get(
            "payment_method_id"
        )

        note = data.get("note", "")

        products = data.get("products", [])

        # ==========================
        # USER
        # ==========================

        user = MUser.objects.filter(
            phone=phone,
            delete_flag=False
        ).first()

        if not user:

            user = MUser.objects.create(
                user_name=user_name,
                phone=phone,
                email="",
                address=""
            )

        # ==========================
        # STATUS
        # ==========================

        status = MStatus.objects.get(
    status_id="01"
)

        # ==========================
        # ORDER HEADER
        # ==========================

        order = TOrderHeader.objects.create(
            user=user,
            user_name=user.user_name,
            order_date=order_date,
            estimated_delivery_date=estimated_delivery,
            total_price_order=total_price,
            payment_method_id=payment_method_id,
            status=status,
            note=note
        )

        # ==========================
        # ORDER DETAIL
        # ==========================

        for item in products:

            product = MProduct.objects.get(
                product_id=item["product_id"]
            )

            TOrderDetail.objects.create(
                order=order,
                product=product,
                product_size=item["size"],
                type=product.type,
                quality=item["quantity"],
                price_product=item["price"]
            )

        return JsonResponse({
            "success": True,
            "order_id": order.order_id
        })

    except Exception as e:

        return JsonResponse({
            "success": False,
            "message": str(e)
        })
@require_POST
def favorite_product(request):

    try:
        data = json.loads(request.body)

        product_id = data.get("product_id")

        user_id = request.session.get("user_id")

        if not user_id:
            return JsonResponse({
                "success": False,
                "error": "Bạn chưa đăng nhập."
            })

        user = MUser.objects.get(
            user_id=user_id,
            delete_flag=False
        )

        product = MProduct.objects.get(
            product_id=product_id,
            delete_flag=False
        )

        favorite, created = TProductFavorite.objects.get_or_create(
            user=user,
            product=product,
            defaults={
                "favorite_flag": True,
                "delete_flag": False
            }
        )

        if not created:
            favorite.favorite_flag = not favorite.favorite_flag
            favorite.save()

        return JsonResponse({
            "success": True,
            "favorite_flag": favorite.favorite_flag
        })

    except Exception as e:
        return JsonResponse({
            "success": False,
            "error": str(e)
        })

def favorite_user(request):
    user_id = request.session.get("user_id")

    favorite_products = MProduct.objects.filter(
        tproductfavorite__user_id=user_id,
        tproductfavorite__favorite_flag=True,
        tproductfavorite__delete_flag=False,
        delete_flag=False
    )

    return render(
        request,
        "user/favorite.html",
        {
            "favorite_products": favorite_products
        }
    )
import json
from django.http import JsonResponse


def save_location(request):
    if request.method == "POST":

        data = json.loads(request.body)

        latitude = data.get("latitude")
        longitude = data.get("longitude")

        user_id = request.session.get("user_id")

        # Lấy địa chỉ từ tọa độ
        url = f"https://nominatim.openstreetmap.org/reverse?format=json&lat={latitude}&lon={longitude}"

        headers = {
            "User-Agent": "HHUONG-Shop"
        }

        response = requests.get(url, headers=headers)

        address = ""

        if response.status_code == 200:
            result = response.json()
            address = result.get("display_name", "")

        if user_id:
            MUser.objects.filter(
                user_id=user_id
            ).update(
                latitude=latitude,
                longitude=longitude,
                address=address
            )

        return JsonResponse({
            "success": True,
            "address": address
        })

    return JsonResponse({
        "success": False
    })